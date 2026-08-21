import openpyxl
import os
import sys

def test_export():
    template_path = r'D:\0 Pre Deploy\Timesheet\Timesheet_Template_v2.xlsx'
    if not os.path.exists(template_path):
        template_path = r'D:\0 Pre Deploy\Timesheet\nextjs-app\public\Timesheet_Template_v2.xlsx'
    
    output_path = r'D:\0 Pre Deploy\Timesheet\nextjs-app\scripts\test_openpyxl_output.xlsx'

    wb = openpyxl.load_workbook(template_path)
    ws = wb['Timesheet']

    # Update Header
    ws['A6'] = 'NAME:  John Doe'
    ws['G6'] = 'Aug-2026'
    ws['G7'] = 'COM116'
    ws['G8'] = 'Commissioning Lead Engineer'

    # Update Day 1 (Row 11)
    ws['D11'] = 8
    ws['E11'] = 2
    ws['F11'] = 0
    ws['G11'] = 'CMN, SAP - Pre-commissioning test'

    # Update Signature
    ws['B57'] = 'John Doe'
    ws['B58'] = 'Commissioning Lead Engineer'

    wb.save(output_path)
    print('Openpyxl successfully written to:', output_path)

if __name__ == '__main__':
    test_export()
