import sys
import json
import os
import datetime
import base64
import io
import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

def generate_metso_timesheet(template_path, output_path, user_id, username, user_role, month_str, records_json, approval_json=None):
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found at: {template_path}")

    year, month_num = map(int, month_str.split('-'))

    records = json.loads(records_json) if isinstance(records_json, str) else records_json
    record_map = {r['date']: r for r in records}

    approval = None
    if approval_json:
        approval = json.loads(approval_json) if isinstance(approval_json, str) else approval_json

    wb = openpyxl.load_workbook(template_path)
    ws = wb['Timesheet']

    # Month short label (e.g. Aug-2026)
    month_names_short = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    month_label = f"{month_names_short[month_num - 1]}-{year}"

    # Header Updates (Preserving styles and fonts)
    ws['A6'] = f"NAME:  {username}"
    ws['G6'] = month_label
    ws['G7'] = user_id
    ws['G8'] = user_role if user_role else 'Commissioning Engineer'

    # Determine total days in month
    if month_num == 12:
        next_month = datetime.date(year + 1, 1, 1)
    else:
        next_month = datetime.date(year, month_num + 1, 1)
    
    total_days = (next_month - datetime.date(year, month_num, 1)).days
    day_names_list = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']

    # Update Daily Rows (11 to 41)
    for day in range(1, 32):
        row_idx = 10 + day  # Row 11 is Day 1

        if day <= total_days:
            dt = datetime.date(year, month_num, day)
            day_name = day_names_list[dt.isoweekday() % 7]  # 7 is Sunday -> 0
            date_str_formatted = dt.strftime('%d/%m/%Y')
            ymd_date = dt.strftime('%Y-%m-%d')
            is_weekend = dt.weekday() in (5, 6) # Sat or Sun

            rec = record_map.get(ymd_date)

            ws[f'A{row_idx}'] = day_name
            ws[f'B{row_idx}'] = date_str_formatted

            if rec:
                reg_hrs = float(rec.get('working_hours') or rec.get('hours') or 0)
                ot_hrs = float(rec.get('overtime_hours') or rec.get('overtime') or 0)
                areas = [rec.get(f'area{i}') for i in range(1, 5) if rec.get(f'area{i}')]
                areas_str = ', '.join(areas)
                remark = rec.get('remark') or ''
                desc = f"{areas_str} - {remark}" if (areas_str and remark) else (areas_str or remark or 'Commissioning Work')

                rmk_upper = remark.upper()
                if 'ROTATION' in rmk_upper:
                    day_type = 'ROTATION'
                elif 'SICK' in rmk_upper:
                    day_type = 'SICK'
                elif 'HOLIDAY' in rmk_upper:
                    day_type = 'HOLIDAY'
                elif 'TRAVEL' in rmk_upper:
                    day_type = 'TRAVEL'
                elif reg_hrs == 0:
                    day_type = 'WEEKLY OFF'
                else:
                    day_type = 'WORK'

                ws[f'C{row_idx}'] = day_type
                ws[f'D{row_idx}'] = reg_hrs
                ws[f'E{row_idx}'] = ot_hrs
                ws[f'F{row_idx}'] = 0
                ws[f'G{row_idx}'] = desc
            else:
                ws[f'C{row_idx}'] = 'WEEKLY OFF' if is_weekend else 'WORK'
                ws[f'D{row_idx}'] = 0
                ws[f'E{row_idx}'] = 0
                ws[f'F{row_idx}'] = 0
                ws[f'G{row_idx}'] = ''
        else:
            # Clear unused row values for shorter months
            ws[f'A{row_idx}'] = ''
            ws[f'B{row_idx}'] = ''
            ws[f'C{row_idx}'] = ''
            ws[f'D{row_idx}'] = 0
            ws[f'E{row_idx}'] = 0
            ws[f'F{row_idx}'] = 0
            ws[f'G{row_idx}'] = ''

    # Signature Block - Employee
    ws['B57'] = username
    ws['B58'] = user_role if user_role else 'Commissioning Engineer'

    # Signature Block - Approver Digital Signature
    sig_tmp_path = None
    if approval and approval.get('signature_data'):
        try:
            sig_data = approval['signature_data']
            if ',' in sig_data:
                sig_data = sig_data.split(',')[1]

            img_bytes = base64.b64decode(sig_data)
            sig_tmp_path = output_path + '.sig.png'

            # Process image to make background 100% transparent using PIL
            pil_img = PILImage.open(io.BytesIO(img_bytes)).convert("RGBA")
            datas = pil_img.getdata()
            newData = []
            for item in datas:
                # Convert white / near-white background pixels to transparent alpha
                if item[0] > 200 and item[1] > 200 and item[2] > 200:
                    newData.append((255, 255, 255, 0))
                else:
                    newData.append(item)
            pil_img.putdata(newData)
            pil_img.save(sig_tmp_path, "PNG")

            # Place transparent signature image at G60 (Under APPROVED BY: SIGNATURE / SITE MANAGER)
            img = Image(sig_tmp_path)
            img.width = 140
            img.height = 55
            ws.add_image(img, 'G60')

            ws['G57'] = approval.get('approver_name') or 'Site Manager'
            ws['G58'] = 'SITE MANAGER / CUSTOMER REP'
        except Exception as e:
            sys.stderr.write(f"Warning: Failed to render approval signature: {e}\n")

    wb.save(output_path)

    if sig_tmp_path and os.path.exists(sig_tmp_path):
        try:
            os.remove(sig_tmp_path)
        except Exception:
            pass

    print(f"SUCCESS:{output_path}")

if __name__ == '__main__':
    if len(sys.argv) < 8:
        print("Usage: python export_metso_template.py <template_path> <output_path> <user_id> <username> <user_role> <month_str> <json_file_path> [approval_json_path]")
        sys.exit(1)

    t_path = sys.argv[1]
    o_path = sys.argv[2]
    u_id = sys.argv[3]
    u_name = sys.argv[4]
    u_role = sys.argv[5]
    m_str = sys.argv[6]
    json_path = sys.argv[7]

    app_json = None
    if len(sys.argv) >= 9 and os.path.exists(sys.argv[8]):
        with open(sys.argv[8], 'r', encoding='utf-8') as f_app:
            app_json = f_app.read()

    with open(json_path, 'r', encoding='utf-8') as f:
        recs = json.load(f)

    generate_metso_timesheet(t_path, o_path, u_id, u_name, u_role, m_str, recs, app_json)
